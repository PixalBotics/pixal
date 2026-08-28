#!/usr/bin/env python3
"""Builds/refreshes the PixalBotics Outreach Tracker spreadsheet from leads.csv."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

LEADS_CSV = "/tmp/pixal/leads/leads.csv"
OUT_XLSX = "/tmp/pixal/leads/PixalBotics_Outreach_Tracker.xlsx"

NAVY = "0B1C3D"
GOLD = "E9B64A"
WHITE = "FFFFFF"
LIGHT = "F2F4F9"

STATUS_OPTIONS = "sourced,audited,emailed,replied,meeting_booked,client,not_interested,no_response"

COLUMNS = [
    ("id", 10), ("name", 26), ("niche", 20), ("country", 12), ("city", 20),
    ("website", 30), ("email", 30), ("phone", 18), ("channel", 22),
    ("site_quality", 30), ("pitch_angle", 40), ("status", 16),
    ("date_emailed", 14), ("date_replied", 14), ("meeting_date", 16),
    ("notes", 30),
]


def load_leads():
    with open(LEADS_CSV, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build():
    leads = load_leads()
    wb = Workbook()
    ws = wb.active
    ws.title = "Outreach Tracker"

    thin = Side(style="thin", color="D8DCE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for i, (col, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=col.replace("_", " ").title())
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Data rows
    for r, lead in enumerate(leads, start=2):
        for c, (col, _) in enumerate(COLUMNS, start=1):
            val = lead.get(col, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in ("pitch_angle", "site_quality", "notes")))
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)

    last_row = len(leads) + 1
    status_col_letter = get_column_letter([c for c, (col, _) in enumerate(COLUMNS, start=1) if col == "status"][0])

    # Status dropdown validation
    dv = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{status_col_letter}2:{status_col_letter}{max(last_row, 200)}")

    # Conditional-ish manual coloring per current status value
    status_colors = {
        "sourced": "D9E2F3",
        "audited": "CFE8FF",
        "emailed": "FFF2CC",
        "replied": "D9EAD3",
        "meeting_booked": "C6E0B4",
        "client": "B7DFB0",
        "not_interested": "F4CCCC",
        "no_response": "EFEFEF",
    }
    status_col_idx = [c for c, (col, _) in enumerate(COLUMNS, start=1) if col == "status"][0]
    for r in range(2, last_row + 1):
        val = ws.cell(row=r, column=status_col_idx).value
        color = status_colors.get((val or "").strip())
        if color:
            ws.cell(row=r, column=status_col_idx).fill = PatternFill("solid", fgColor=color)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "PixalBotics Outreach Summary"
    ws2["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws2["A3"] = "Total Leads"
    ws2["B3"] = f"=COUNTA('Outreach Tracker'!A2:A{max(last_row,200)})"
    statuses = ["sourced", "audited", "emailed", "replied", "meeting_booked", "client", "not_interested", "no_response"]
    for i, s in enumerate(statuses, start=4):
        ws2.cell(row=i, column=1, value=s.replace("_", " ").title())
        ws2.cell(row=i, column=2, value=f"=COUNTIF('Outreach Tracker'!{status_col_letter}2:{status_col_letter}{max(last_row,200)},\"{s}\")")
    for col, width in [("A", 22), ("B", 12)]:
        ws2.column_dimensions[col].width = width

    wb.save(OUT_XLSX)
    print(f"Saved {OUT_XLSX} with {len(leads)} leads.")


if __name__ == "__main__":
    build()
